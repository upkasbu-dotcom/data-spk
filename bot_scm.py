import time
import requests
import json
import os
import io
import sys
import traceback
import shutil

from github import Github
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook


def upload_to_github(data, filename="data_scm.json"):
    """
    Upload data JSON ke repository GitHub.
    Jika file sudah ada, akan di-update. Jika belum, akan dibuat baru.
    """
    token = os.environ.get('GITHUB_TOKEN')
    if not token:
        print("[ERROR] GITHUB_TOKEN tidak ditemukan di environment variables!")
        return False

    try:
        g = Github(token)
        repo = g.get_repo("upkasbu-dotcom/data-spk")
        file_content = json.dumps(data, indent=4, ensure_ascii=False)

        try:
            contents = repo.get_contents(filename)
            repo.update_file(
                path=contents.path,
                message="Update data SCM otomatis via GitHub Actions",
                content=file_content,
                sha=contents.sha
            )
            print(f"[OK] File {filename} berhasil diperbarui di GitHub.")
        except Exception:
            repo.create_file(
                path=filename,
                message="Commit data SCM pertama",
                content=file_content
            )
            print(f"[OK] File {filename} berhasil dibuat di GitHub.")

        return True

    except Exception as e:
        print(f"[ERROR] Gagal upload ke GitHub: {e}")
        traceback.print_exc()
        return False


def create_chrome_driver():
    """
    Membuat instance Chrome WebDriver yang kompatibel dengan GitHub Actions.
    """
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-software-rasterizer')
    options.add_argument('--remote-allow-origins=*')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument(
        '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    )
    # Hilangkan flag 'webdriver' agar tidak mudah terdeteksi
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    chromedriver_path = shutil.which('chromedriver')
    if chromedriver_path:
        print(f"[INFO] ChromeDriver ditemukan di: {chromedriver_path}")
        service = Service(executable_path=chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)
    else:
        print("[WARNING] ChromeDriver tidak ditemukan di PATH, menggunakan default...")
        driver = webdriver.Chrome(options=options)

    driver.set_page_load_timeout(90)

    # Coba sembunyikan property navigator.webdriver
    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            },
        )
    except Exception:
        pass

    return driver


def _login_berhasil(driver, wait):
    """
    Cek apakah login berhasil menggunakan beberapa indikator.
    Menggantikan EC.any_of yang hanya ada di Selenium >= 4.5.
    """
    indikator_url = ["dashboard", "home", "surat-perintah-kerja", "spk", "admin"]
    indikator_xpath = [
        "//nav",
        "//header",
        "//aside",
        "//sidebar",
        "//button[contains(text(),'Logout')]",
        "//button[contains(text(),'Log out')]",
        "//a[contains(@href,'logout')]",
        "//a[contains(@href,'profil')]",
    ]

    deadline = time.time() + 30
    while time.time() < deadline:
        current_url = driver.current_url.lower()
        if any(keyword in current_url for keyword in indikator_url):
            return True

        for xpath in indikator_xpath:
            try:
                elem = driver.find_element(By.XPATH, xpath)
                if elem.is_displayed():
                    return True
            except Exception:
                continue

        time.sleep(1)

    return False


def jalankan_bot():
    """
    Fungsi utama bot SCM Nusadaya.
    """
    print("=" * 50)
    print("=== [START] Operasi Bot SCM ===")
    print("=" * 50)

    USERNAME = os.environ.get('USERNAME')
    PASSWORD = os.environ.get('PASSWORD')

    if not USERNAME or not PASSWORD:
        print("[ERROR] USERNAME atau PASSWORD tidak ditemukan di environment variables!")
        print("[ERROR] Pastikan secrets sudah dikonfigurasi di repository Settings > Secrets.")
        return False

    driver = create_chrome_driver()

    try:
        # ===== STEP 1: Login =====
        print("[STEP 1] Membuka halaman login...")
        driver.get("https://scm.nusadaya.net/login")
        wait = WebDriverWait(driver, 30)

        # Tunggu sampai minimal ada satu input terlihat
        print("[STEP 1] Menunggu form login muncul...")
        wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[@type='text' or @type='email' or @placeholder]")
            )
        )

        # Cari input username/email (case-insensitive via translate)
        print("[STEP 1] Mengisi USERNAME...")
        email_input = None
        email_xpaths = [
            "//input[translate(@type,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='text']",
            "//input[translate(@type,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='email']",
            "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
            "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'nip')]",
            "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'username')]",
            "//input[@name='email' or @name='username' or @name='nip']",
        ]
        for xp in email_xpaths:
            try:
                email_input = driver.find_element(By.XPATH, xp)
                if email_input.is_displayed():
                    break
            except Exception:
                email_input = None

        if email_input is None:
            raise Exception("Input username/email tidak ditemukan!")

        email_input.clear()
        email_input.send_keys(USERNAME)

        # Cari input password (FIX: case-sensitive, harus huruf kecil 'password')
        print("[STEP 1] Mengisi PASSWORD...")
        password_input = None
        password_xpaths = [
            "//input[translate(@type,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz')='password']",
            "//input[@type='password']",
            "//input[@name='password']",
            "//input[@autocomplete='current-password']",
        ]
        for xp in password_xpaths:
            try:
                password_input = driver.find_element(By.XPATH, xp)
                if password_input.is_displayed():
                    break
            except Exception:
                password_input = None

        if password_input is None:
            raise Exception("Input password tidak ditemukan!")

        password_input.clear()
        password_input.send_keys(PASSWORD)

        # Klik tombol login (coba beberapa variasi teks)
        print("[STEP 1] Klik tombol Login...")
        login_button = None
        login_xpaths = [
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login')]",
            "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'masuk')]",
            "//button[@type='submit']",
            "//input[@type='submit']",
            "//button[contains(@class,'login')]",
        ]
        for xp in login_xpaths:
            try:
                login_button = driver.find_element(By.XPATH, xp)
                if login_button.is_displayed() and login_button.is_enabled():
                    break
            except Exception:
                login_button = None

        if login_button is None:
            raise Exception("Tombol login tidak ditemukan!")

        login_button.click()

        print("[STEP 1] Menunggu redirect setelah login...")
        time.sleep(3)

        if _login_berhasil(driver, wait):
            print(f"[STEP 1] Login berhasil! URL: {driver.current_url}")
        else:
            print("[WARNING] Tidak dapat memverifikasi login berhasil, melanjutkan...")
            print(f"[WARNING] URL saat ini: {driver.current_url}")

        # ===== STEP 2: Download Excel =====
        print("[STEP 2] Mengambil session cookies...")
        session_cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        print(f"[STEP 2] Berhasil mengambil {len(session_cookies)} cookies.")

        # Tambahkan XSRF-TOKEN dan session cookie bila ada
        headers = {
            'Referer': 'https://scm.nusadaya.net/surat-perintah-kerja',
            'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/octet-stream, */*',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        download_url = "https://scm.nusadaya.net/surat-perintah-kerja/export"
        print(f"[STEP 2] Mendownload dari: {download_url}")

        max_retries = 3
        response_dl = None

        for attempt in range(1, max_retries + 1):
            try:
                response_dl = requests.get(
                    download_url,
                    cookies=session_cookies,
                    headers=headers,
                    timeout=60,
                    allow_redirects=True
                )
                if response_dl.status_code == 200:
                    print(f"[STEP 2] Download berhasil! (Attempt {attempt})")
                    print(f"[STEP 2] Content-Type: {response_dl.headers.get('Content-Type')}")
                    print(f"[STEP 2] Content-Length: {len(response_dl.content)} bytes")
                    break
                else:
                    print(f"[STEP 2] Status code: {response_dl.status_code} (Attempt {attempt}/{max_retries})")
                    if response_dl.status_code in (401, 403):
                        print("[STEP 2] Akses ditolak, kemungkinan session tidak valid.")
            except requests.exceptions.RequestException as e:
                print(f"[STEP 2] Request error: {e} (Attempt {attempt}/{max_retries})")

            if attempt < max_retries:
                wait_time = attempt * 5
                print(f"[STEP 2] Menunggu {wait_time} detik sebelum retry...")
                time.sleep(wait_time)

        # ===== STEP 3: Proses & Upload =====
        if response_dl and response_dl.status_code == 200:
            print("[STEP 3] Memproses file Excel...")
            try:
                wb = load_workbook(filename=io.BytesIO(response_dl.content), data_only=True)
                all_data = {}

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        rows.append([cell if cell is not None else "" for cell in row])
                    all_data[sheet_name] = rows

                print(f"[STEP 3] Berhasil membaca {len(all_data)} sheet: {list(all_data.keys())}")

                print("[STEP 3] Mengupload ke GitHub...")
                success = upload_to_github(all_data)

                if success:
                    print("[STEP 3] Data berhasil diupload ke GitHub!")
                else:
                    print("[STEP 3] Gagal mengupload ke GitHub!")

            except Exception as e:
                print(f"[ERROR] Gagal memproses file Excel: {e}")
                traceback.print_exc()
        else:
            status = response_dl.status_code if response_dl else "No response"
            print(f"[ERROR] Gagal download file! Status: {status}")
            return False

        print("=" * 50)
        print("=== [FINISH] Operasi Selesai ===")
        print("=" * 50)
        return True

    except Exception as e:
        print(f"[FATAL ERROR] Terjadi error: {e}")
        traceback.print_exc()
        # Simpan screenshot untuk debugging
        try:
            driver.save_screenshot("error_screenshot.png")
            print("[DEBUG] Screenshot disimpan: error_screenshot.png")
            with open("error_page.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("[DEBUG] Page source disimpan: error_page.html")
        except Exception:
            pass
        return False

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print("[INFO] Browser ditutup.")


if __name__ == "__main__":
    result = jalankan_bot()
    sys.exit(0 if result else 1)
